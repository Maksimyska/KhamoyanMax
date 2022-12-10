import csv
import re
import os
from functools import cmp_to_key
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
import matplotlib.pyplot as plt
import numpy as np
from jinja2 import Environment, FileSystemLoader
import pdfkit


class Vacancy:
    def __init__(self, name, description, key_skills, experience_id, premium,
                 employer_name, salary, area_name, published_at):
        self.name = name
        self.description = description
        self.key_skills = key_skills
        self.experience_id = experience_id
        self.premium = premium
        self.employer_name = employer_name
        self.salary = salary
        self.area_name = area_name
        self.published_at = published_at


class Salary:
    def __init__(self, salary_from, salary_to, salary_gross, salary_currency):
        self.salary_from = salary_from
        self.salary_to = salary_to
        self.salary_gross = salary_gross
        self.salary_currency = salary_currency

    def currency_to_rur(self):
        currency_to_rub = {
            "Манаты": 35.68,
            "Белорусские рубли": 23.91,
            "Евро": 59.90,
            "Грузинский лари": 21.74,
            "Киргизский сом": 0.76,
            "Тенге": 0.13,
            "Рубли": 1,
            "Гривны": 1.64,
            "Доллары": 60.66,
            "Узбекский сум": 0.0055
        }

        return list(map(lambda x: int(x.replace(' ', '')) * currency_to_rub[self.salary_currency],
                        (self.salary_from, self.salary_to)))

    def get_salary(self):
        return sum(self.currency_to_rur()) / 2


class DataSet:
    def __init__(self, file_name):
        self.file_name = file_name
        (headers, info) = self.csv_reader()
        vacancies = self.csv_filter(headers, info) if len(
            headers) > 6 else self.mini_csv_filter(info)
        self.vacancies_objects = vacancies

    def csv_reader(self):
        with open(self.file_name, encoding="utf-8-sig") as file:
            reader = [x for x in csv.reader(file)]
            title = reader.pop(0)
            titleCount = len(title)
            info = list(filter(lambda data: '' not in data and len(
                data) == titleCount, reader))
        return title, info

    @staticmethod
    def mini_csv_filter(info):
        vacancies = []
        for item in info:
            salary = Salary(item[1], item[2], None, item[3])
            vacancies.append(
                Vacancy(item[0], None, None, None, None, None, salary, item[4], item[5]))
        return vacancies

    @staticmethod
    def csv_filter(title, info):
        def normalize_csv_file(info_cell):
            temp_info = "__temp__".join(info_cell.split("\n"))
            temp_info = re.sub(r"<[^<>]*>", "", temp_info)
            temp_info = re.sub(r"\s+", " ", temp_info)
            return str.strip(temp_info)

        vacancies = []
        for item in info:
            info_list = list(
                map(lambda x: normalize_csv_file(item[x]), range(len(title))))
            salary = Salary(info_list[6], info_list[7],
                            info_list[8], info_list[9])
            key_skills = info_list[2].split('__temp__')
            vacancy = Vacancy(
                info_list[0],
                info_list[1],
                key_skills,
                info_list[3],
                info_list[4],
                info_list[5],
                salary,
                info_list[10],
                info_list[11]
            )
            vacancies.append(vacancy)
        return vacancies


class InputConnect:
    @staticmethod
    def info_formatter(vacancies):
        def formatter_string_number(str_num):
            return str_num if str_num.find('.') == -1 else str_num[:len(str_num) - 2]

        def formatter_salary(attr_value):
            salary_from = formatter_string_number(attr_value.salary_from)
            salary_to = formatter_string_number(attr_value.salary_to)
            salary_currency = dic_currency[attr_value.salary_currency]
            return Salary(salary_from, salary_to, None, salary_currency)

        def formatter_published_at(attr_value):
            return attr_value[0:4]

        dic_currency = {
            "AZN": "Манаты",
            "BYR": "Белорусские рубли",
            "EUR": "Евро",
            "GEL": "Грузинский лари",
            "KGS": "Киргизский сом",
            "KZT": "Тенге",
            "RUR": "Рубли",
            "UAH": "Гривны",
            "USD": "Доллары",
            "UZS": "Узбекский сум"
        }

        for vacancy in vacancies:
            setattr(vacancy, "salary", formatter_salary(
                getattr(vacancy, "salary")))
            setattr(vacancy, "published_at", formatter_published_at(
                getattr(vacancy, "published_at")))
        return vacancies

    def info_finder(self, vacancies, parameter):
        salary_level_by_years, selected_vacancy_salary_year, count_vacancies_by_year, selected_vacancy_year_count, \
            salary_levels_by_city, count_vacancies_by_city = {}, {}, {}, {}, {}, {}
        for item in vacancies:
            salary = item.salary.get_salary()
            year = int(item.published_at)
            if year not in salary_level_by_years:
                salary_level_by_years[year] = (salary, 1)
                count_vacancies_by_year[year] = 1
                selected_vacancy_salary_year[year] = (0, 0)
                selected_vacancy_year_count[year] = 0
            else:
                sal_lvl_by_y = salary_level_by_years[year]
                salary_level_by_years[year] = (
                    sal_lvl_by_y[0] + salary, sal_lvl_by_y[1] + 1)
                count_vacancies_by_year[year] += 1
            if parameter in item.name:
                sel_vac_sal_y = selected_vacancy_salary_year[year]
                selected_vacancy_salary_year[year] = (
                    sel_vac_sal_y[0] + salary, sel_vac_sal_y[1] + 1)
                selected_vacancy_year_count[year] += 1
            if item.area_name not in salary_levels_by_city:
                count_vacancies_by_city[item.area_name] = 1
                salary_levels_by_city[item.area_name] = (salary, 1)
            else:
                sal_lvl_by_c = salary_levels_by_city[item.area_name]
                salary_levels_by_city[item.area_name] = (
                    sal_lvl_by_c[0] + salary, sal_lvl_by_c[1] + 1)
                count_vacancies_by_city[item.area_name] += 1
        return self.info_calculating(salary_level_by_years, selected_vacancy_salary_year, count_vacancies_by_year,
                                     selected_vacancy_year_count, salary_levels_by_city, count_vacancies_by_city,
                                     len(vacancies))

    @staticmethod
    def info_calculating(salary_level_by_years, selected_vacancy_salary_year, count_vacancies_by_year,
                         selected_vacancy_year_count, salary_levels_by_city, count_vacancies_by_city, vacancies_count):

        def sort(dictionary):
            dict_item = [(key, value) for key, value in dictionary.items()]
            dict_item.sort(key=cmp_to_key(
                lambda x, y: -1 if x[1] <= y[1] else 1))
            return dict(dict_item)

        (salary_level_by_years, selected_vacancy_salary_year, salary_levels_by_city) = \
            list(map(lambda dictionary:
                     dict(map(lambda dict_item:
                              (dict_item[0], int(dict_item[1][0] / dict_item[1][1]) if dict_item[1][1] != 0
                               else int(dict_item[1][0])), dictionary.items())),
                     (salary_level_by_years, selected_vacancy_salary_year, salary_levels_by_city)))
        count_vacancies_by_city = dict(map(lambda dict_pair: (dict_pair[0],
                                                              float(f"{dict_pair[1] / vacancies_count:.4f}")),
                                           count_vacancies_by_city.items()))
        count_vacancies_by_city = dict(
            filter(lambda dict_pair: dict_pair[1] >= 0.01, count_vacancies_by_city.items()))
        count_vacancies_by_city = sort(count_vacancies_by_city)
        count_vacancies_by_city = {k: count_vacancies_by_city[k] for k in list(
            count_vacancies_by_city)[-10:][::-1]}
        count_vacancies_by_city = dict(map(lambda dict_pair: (dict_pair[0], f"{round(dict_pair[1] * 100, 2)}%"),
                                           count_vacancies_by_city.items()))
        salary_levels_by_city = dict(filter(lambda dict_pair: dict_pair[0] in count_vacancies_by_city,
                                            salary_levels_by_city.items()))
        salary_levels_by_city = sort(salary_levels_by_city)
        salary_levels_by_city = {k: salary_levels_by_city[k] for k in list(
            salary_levels_by_city)[-10:][::-1]}
        return salary_level_by_years, selected_vacancy_salary_year, count_vacancies_by_year, \
            selected_vacancy_year_count, salary_levels_by_city, count_vacancies_by_city, vacancies_count


class Report:
    def __init__(self, vacancy_info):
        self.salaries_year_level = vacancy_info[0]
        self.vacancies_year_count = vacancy_info[1]
        self.selected_salary_year_level = vacancy_info[2]
        self.selected_vacancy_year_count = vacancy_info[3]
        self.salaries_city_level = vacancy_info[4]
        self.vacancies_city_count = vacancy_info[5]

    def generate_excel(self, vacancy_name):
        workbook = Workbook()
        stats_by_year = workbook.worksheets[0]
        stats_by_year.title = "Cтатистика по годам"
        stats_by_city = workbook.create_sheet("Cтатистика по городам")

        stats_by_year.append(["Год", "Средняя зарплата", f"Средняя зарплата - {vacancy_name}",
                              "Количество вакансий", f"Количество вакансий - {vacancy_name}"])
        for i, year in enumerate(self.salaries_year_level.keys(), 2):
            stats_by_year.cell(row=i, column=1, value=year)
            for j, dictionary in enumerate((self.salaries_year_level, self.vacancies_year_count,
                                            self.selected_salary_year_level, self.selected_vacancy_year_count), 2):
                stats_by_year.cell(row=i, column=j, value=dictionary[year])

        stats_by_city.append(
            ["Город", "Уровень зарплат", "", "Город", "Доля вакансий"])
        for i, city in enumerate(self.salaries_city_level.keys(), 2):
            stats_by_city.cell(row=i, column=1, value=city)
            stats_by_city.cell(
                row=i, column=2, value=self.salaries_city_level[city])
        for i, city in enumerate(self.vacancies_city_count.keys(), 2):
            stats_by_city.cell(row=i, column=4, value=city)
            stats_by_city.cell(
                row=i, column=5, value=self.vacancies_city_count[city])

        self.workbook(workbook)
        workbook.save('report.xlsx')

    @staticmethod
    def workbook(wb):
        bold_font = Font(bold=True)
        thin = Side(border_style="thin", color="000000")
        outline = Border(top=thin, left=thin, right=thin, bottom=thin)
        for worksheet in wb.worksheets:
            for column_cells in worksheet.columns:
                length = max(len(str(cell.value) if cell.value is not None else "")
                             for cell in column_cells)
                worksheet.column_dimensions[column_cells[0]
                                            .column_letter].width = length + 3
            for cell in worksheet[1]:
                cell.font = bold_font
            for column in tuple(worksheet.columns):
                if column[1].value is None:
                    continue
                for cell in column:
                    cell.border = outline

    def generate_image(self, vacancy_name):
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(
            2, 2, figsize=(12, 7.5), layout='constrained')
        self.generate_salary_year_levels_graph(ax1, vacancy_name)
        self.generate_vacancy_year_count_graph(ax2, vacancy_name)
        self.generate_salary_city_levels_graph(ax3)
        self.generate_vacancy_city_count_graph(ax4)
        plt.savefig('graph.png')

    def generate_salary_year_levels_graph(self, ax, vacancy_name):
        ax_labels = self.salaries_year_level.keys()
        x = np.arange(len(ax_labels))
        width = 0.35
        ax.bar(x - width / 2, self.salaries_year_level.values(),
               width, label='Средняя з/п')
        ax.bar(x + width / 2, self.selected_salary_year_level.values(),
               width, label=f'З/п {vacancy_name}')
        ax.set_xticks(x, ax_labels, fontsize=8, rotation=90, ha='right')
        ax.set_title("Уровень зарплат по годам")
        ax.yaxis.grid(True)
        ax.legend(fontsize=8, loc='upper left')

    def generate_vacancy_year_count_graph(self, ax, vacancy_name):
        ax_labels = self.vacancies_year_count.keys()
        x = np.arange(len(ax_labels))
        width = 0.35
        ax.bar(x - width / 2, self.vacancies_year_count.values(),
               width, label='Количество вакансий')
        ax.bar(x + width / 2, self.selected_vacancy_year_count.values(),
               label=f'Количество вакансий {vacancy_name}')
        ax.set_xticks(x, ax_labels, fontsize=8, rotation=90, ha='right')
        ax.set_title("Количество вакансий по годам")
        ax.yaxis.grid(True)
        ax.legend(fontsize=8, loc='upper left')

    def generate_salary_city_levels_graph(self, ax):
        ax_labels = self.salaries_city_level.keys()
        y_pos = np.arange(len(ax_labels))
        ax.barh(y_pos, self.salaries_city_level.values(), align='center')
        ax.set_yticks(y_pos, fontsize=8, labels=ax_labels)
        ax.invert_yaxis()
        ax.set_title("Уровень зарплат по городам")

    def generate_vacancy_city_count_graph(self, ax):
        ax_labels, values = list(
            self.vacancies_city_count.keys()), self.vacancies_city_count.values()
        ax_labels.append('Другие')
        values = list(map(lambda value: float(value[:-1]), values))
        values.append(100 - sum(values))
        ax.pie(values, labels=ax_labels)
        ax.set_title("Доля вакансий по городам")

    def generate_pdf(self, vacancy_name):
        headers1, headers2, headers3 = (["Год", "Средняя зарплата", f"Средняя зарплата - {vacancy_name}",
                                        "Количество вакансий", f"Количество вакансий - {vacancy_name}"],
                                        ["Город", "Уровень зарплат"], ["Город", "Доля вакансий"])
        rows1 = list(map(lambda year: [year] + [dictionary[year] for dictionary in
                                                (self.salaries_year_level, self.vacancies_year_count,
                                                 self.selected_salary_year_level, self.selected_vacancy_year_count)], self.salaries_year_level.keys()))
        rows2 = list(map(lambda city: [
                     city, self.salaries_city_level[city]], self.salaries_city_level.keys()))
        rows3 = list(map(lambda city: [
                     city, self.vacancies_city_count[city]], self.vacancies_city_count.keys()))

        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template.html")
        pdf_template = template.render(graph_name='graph.png',
                                       vacancy_name=vacancy_name, headers1=headers1, headers2=headers2, headers3=headers3,
                                       rows1=rows1, rows2=rows2, rows3=rows3)
        config = pdfkit.configuration(
            wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        options = {'enable-local-file-access': None}
        pdfkit.from_string(pdf_template, 'report.pdf',
                           options=options, configuration=config)


def normalize_input_info(input_info):
    if os.stat(input_info[0]).st_size == 0:
        return "Пустой файл"
    return "Нормализация прошла успешно"


def get_info():
    input_requests = ["Введите название файла: ",
                      "Введите название профессии: "]
    input_info = ["vacancies_by_year.csv", "Javascript"]
    # input_info = [input(input_request) for input_request in input_requests]
    if os.stat(input_info[0]).st_size == 0:
        print("Пустой файл")
        return
    data_set = DataSet(input_info[0])
    if len(data_set.vacancies_objects) == 0:
        print("Нет данных")
        return
    input_connect = InputConnect()
    formatted_info = input_connect.info_formatter(data_set.vacancies_objects)
    info = input_connect.info_finder(formatted_info, input_info[1])
    report = Report(info)
    report.generate_excel(input_info[1])
    report.generate_image(input_info[1])
    report.generate_pdf(input_info[1])


get_info()