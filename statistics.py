import csv
import re
import os
from functools import cmp_to_key
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
import matplotlib.pyplot as plt
import numpy as np
from jinja2 import Environment, FileSystemLoader
import pdfkit


class Vacancy:
    """Класс Вакансий

    Attributes:
        name (str): Имя вакансии
        description (str):  Описание вакансии
        key_skills (list[str]): Скиллы, необходимые вакансии
        experience_id (str): Опыт, необходимый для вакансии
        premium (str): Премиум вакансия или нет
        employer_name (str): Имя компании
        salary (Salary): Оклад
        area_name (str): Местонахождение
        published_at (str): Время публикации вакансии
    """

    def __init__(self, name, description, key_skills, experience_id, premium,
                 employer_name, salary, area_name, published_at):
        """Инициализирует объект Vacancy

        Args:
            name (str):
            description (str | None):
            key_skills (list[str] | None):
            experience_id (str | None):
            premium (str | None):
            employer_name (str | None):
            salary (Salary):
            area_name (str):
            published_at (str):
        """
        self.name = name
        self.description = description
        self.key_skills = key_skills
        self.experience_id = experience_id
        self.premium = premium
        self.employer_name = employer_name
        self.salary = salary
        self.area_name = area_name
        self.published_at = published_at


class Salary:
    """Класс для представления оклада

    Attributes:
        salary_from (str):
        salary_to (str):
        salary_gross (str):
        salary_currency (str):
    """

    def __init__(self, salary_from, salary_to, salary_gross, salary_currency):
        """Инициализирует объект Salary

        Args:
            salary_from (str):
            salary_to (str):
            salary_gross (any):
            salary_currency (str):
        """
        self.salary_from = salary_from
        self.salary_to = salary_to
        self.salary_gross = salary_gross
        self.salary_currency = salary_currency

    def currency_to_rur(self):
        """Перевод в верхнии и нижние вилки оклада

        Returns:
            list[int,int]: Верхняя и нижняя вилки оклада в рублях
        """
        currency_to_rub = {
            "Манаты": 35.68,
            "Белорусские рубли": 23.91,
            "Евро": 59.90,
            "Грузинский лари": 21.74,
            "Киргизский сом": 0.76,
            "Тенге": 0.13,
            "Рубли": 1,
            "Гривны": 1.64,
            "Доллары": 60.66,
            "Узбекский сум": 0.0055
        }

        return list(map(lambda x: int(x.replace(' ', '')) * currency_to_rub[self.salary_currency],
                        (self.salary_from, self.salary_to)))

    def get_salary(self):
        """Получение среднего значения оклада

        Returns:
            float: Среднее значение оклада
        """
        return sum(self.currency_to_rur()) / 2


class DataSet:
    """Класс обработчик csv файла

    Attributes:
        file_name (str): Имя csv файла
        vacancies_objects (list[Vacancy]): Список вакансий полученных из csv файла
    """

    def __init__(self, file_name):
        """Инициализирует объект DataSet

        """
        self.file_name = file_name
        (headers, info) = self.csv_reader()
        vacancies = self.csv_filter(headers, info) if len(headers) > 6 else self.mini_csv_filter(info)
        self.vacancies_objects = vacancies

    def csv_reader(self):
        """Чтение csv файла.

        Args:
            file_name (str): Название csv файла
        Returns:
            title, info: Результат чтения из csv файла в виде пары: лист с названиями столбцов, лист с основными данными
        """
        with open(self.file_name, encoding="utf-8-sig") as file:
            reader = [x for x in csv.reader(file)]
            title = reader.pop(0)
            titleCount = len(title)
            info = list(filter(lambda data: '' not in data and len(data) == titleCount, reader))
        return title, info

    @staticmethod
    def csv_filter(headers, info):
        """Преобразование данных из csv файла в список вакансий, в котором каждой вакансии соответствует одна строка
            из файла
        Args:
            headers (list[str]): Названия столбцов
            info (list[list[str]]): Основные данные csv файла
        Returns:
            list[Vacancy]: Форматированный список вакансий
        """

        def normalize_csv_file(info_cell):
            """Удаление лишних символов из строки для записи в объект Вакансии (html-тегов и т.д.)
            Args:
                info_cell (str): Строка для нормализации
            Returns:
                str: Нормализованная строка
            """
            temp_info = "__temp__".join(info_cell.split("\n"))
            temp_info = re.sub(r"<[^<>]*>", "", temp_info)
            temp_info = re.sub(r"\s+", " ", temp_info)
            return str.strip(temp_info)

        vacancies = []
        for info_row in info:
            info_list = list(map(lambda x: normalize_csv_file(info_row[x]), range(len(headers))))
            salary = Salary(info_list[6], info_list[7], info_list[8], info_list[9])
            key_skills = info_list[2].split('__temp__')
            vacancy = Vacancy(info_list[0], info_list[1], key_skills, info_list[3], info_list[4],
                              info_list[5], salary, info_list[10], info_list[11])
            vacancies.append(vacancy)
        return vacancies

    @staticmethod
    def mini_csv_filter(info):
        vacancies = []
        for item in info:
            salary = Salary(item[1], item[2], None, item[3])
            vacancies.append(
                Vacancy(item[0], None, None, None, None, None, salary, item[4], item[5]))
        return vacancies


class InputConnect:
    """Класс для работы над списком Vacancy

    """

    @staticmethod
    def info_formatter(vacancies):
        """Нормализация данных в вакансиях

        Args:
            vacancies (list[Vacancy] | Vacancy): Список вакансий

        Returns:
            list[Vacancy] | Vacancy: Результат форматирования
        """

        def formatter_string_number(str_num):
            """Устранение дробных разделителей в строковом числе

            Args:
                str_num (str): Число для нормализации

            Returns:
                str: Результат форматирования числа
            """
            return str_num if str_num.find('.') == -1 else str_num[:len(str_num) - 2]

        def formatter_salary(attr_value):
            """Преобразование оклада
            Args:
                attr_value (Salary): Объект оклада
            Returns:
                Salary: Результат форматирования оклада
            """
            salary_from = formatter_string_number(attr_value.salary_from)
            salary_to = formatter_string_number(attr_value.salary_to)
            salary_currency = dic_currency[attr_value.salary_currency]
            return Salary(salary_from, salary_to, None, salary_currency)

        def formatter_published_at(attr_value):
            """Получение года из строки, содержащей дату

            """
            return attr_value[0:4]

        dic_currency = {
            "AZN": "Манаты",
            "BYR": "Белорусские рубли",
            "EUR": "Евро",
            "GEL": "Грузинский лари",
            "KGS": "Киргизский сом",
            "KZT": "Тенге",
            "RUR": "Рубли",
            "UAH": "Гривны",
            "USD": "Доллары",
            "UZS": "Узбекский сум"
        }

        for vacancy in vacancies:
            setattr(vacancy, "salary", formatter_salary(getattr(vacancy, "salary")))
            setattr(vacancy, "published_at", formatter_published_at(getattr(vacancy, "published_at")))
        return vacancies

    def info_finder(self, vacancies, parameter):
        """Формирование информации по годам о вакансиях:
            уровень зарплат по годам, уровень зарплат по годам для выбранной вакансии,
            количество вакансий по годам,
            количество вакансий по годам для выбранной вакансии,
            уровень зарплат по городам,
            количество вакансий по городам,
            общее количество вакансий
        """
        salary_level_by_years, selected_vacancy_salary_year, count_vacancies_by_year, selected_vacancy_year_count, \
        salary_levels_by_city, count_vacancies_by_city = {}, {}, {}, {}, {}, {}
        for item in vacancies:
            salary = item.salary.get_salary()
            year = int(item.published_at)
            if year not in salary_level_by_years:
                salary_level_by_years[year] = (salary, 1)
                count_vacancies_by_year[year] = 1
                selected_vacancy_salary_year[year] = (0, 0)
                selected_vacancy_year_count[year] = 0
            else:
                sal_yr_lvl = salary_level_by_years[year]
                salary_level_by_years[year] = (sal_yr_lvl[0] + salary, sal_yr_lvl[1] + 1)
                count_vacancies_by_year[year] += 1
            if parameter in item.name:
                sel_sal_ye_lvl = selected_vacancy_salary_year[year]
                selected_vacancy_salary_year[year] = (sel_sal_ye_lvl[0] + salary, sel_sal_ye_lvl[1] + 1)
                selected_vacancy_year_count[year] += 1
            if item.area_name not in salary_levels_by_city:
                count_vacancies_by_city[item.area_name] = 1
                salary_levels_by_city[item.area_name] = (salary, 1)
            else:
                sal_ct_lvl = salary_levels_by_city[item.area_name]
                salary_levels_by_city[item.area_name] = (sal_ct_lvl[0] + salary, sal_ct_lvl[1] + 1)
                count_vacancies_by_city[item.area_name] += 1
        return self.info_calculating(salary_level_by_years, selected_vacancy_salary_year, count_vacancies_by_year,
                                     selected_vacancy_year_count, salary_levels_by_city, count_vacancies_by_city,
                                     len(vacancies))

    @staticmethod
    def info_calculating(salary_level_by_years, selected_vacancy_salary_year, count_vacancies_by_year,
                         selected_vacancy_year_count, salary_levels_by_city, count_vacancies_by_city, vacancies_count):
        """Окончательное форматирование словарей, фильтрация, сортировка, выборка первого десятка для некоторых

        """

        def sort(dictionary):
            """Сортировка словаря лексикографически
            Args:
                dictionary (dict[str: int]): Словарь для сортировки
            Returns:
                dict[str: int]: Отсортированный словарь
            """
            dict_pairs = [(key, value) for key, value in dictionary.items()]
            dict_pairs.sort(key=cmp_to_key(lambda x, y: -1 if x[1] <= y[1] else 1))
            return dict(dict_pairs)

        (salary_level_by_years, selected_vacancy_salary_year, salary_levels_by_city) = \
            list(map(lambda dictionary:
                     dict(map(lambda dict_pair:
                              (dict_pair[0], int(dict_pair[1][0] / dict_pair[1][1]) if dict_pair[1][1] != 0
                              else int(dict_pair[1][0])), dictionary.items())),
                     (salary_level_by_years, selected_vacancy_salary_year, salary_levels_by_city)))
        count_vacancies_by_city = dict(map(lambda dict_pair: (dict_pair[0],
                                                              float(f"{dict_pair[1] / vacancies_count:.4f}")),
                                           count_vacancies_by_city.items()))
        count_vacancies_by_city = dict(filter(lambda dict_pair: dict_pair[1] >= 0.01, count_vacancies_by_city.items()))
        count_vacancies_by_city = sort(count_vacancies_by_city)
        count_vacancies_by_city = {k: count_vacancies_by_city[k] for k in list(count_vacancies_by_city)[-10:][::-1]}
        count_vacancies_by_city = dict(map(lambda dict_pair: (dict_pair[0], f"{round(dict_pair[1] * 100, 2)}%"),
                                           count_vacancies_by_city.items()))
        salary_levels_by_city = dict(filter(lambda dict_pair: dict_pair[0] in count_vacancies_by_city,
                                            salary_levels_by_city.items()))
        salary_levels_by_city = sort(salary_levels_by_city)
        salary_levels_by_city = {k: salary_levels_by_city[k] for k in list(salary_levels_by_city)[-10:][::-1]}
        return salary_level_by_years, selected_vacancy_salary_year, count_vacancies_by_year, \
               selected_vacancy_year_count, salary_levels_by_city, count_vacancies_by_city, vacancies_count


class Report:
    """Класс для генерации файлов по анализу статистики: графиков, excel таблиц, общего pdf-файла

    Attributes:
        salaries_year_level (dict[int: tuple[int, int]]): ЗП по годам
        selected_salary_year_level (dict[int: tuple[int, int]]): Уровень зарплат по годам для выбранной вакансии
        vacancies_year_count (dict[int: int]): Количество вакансий по годам
        selected_vacancy_year_count (dict[int: int]): Количество вакансий по годам для выбранной вакансии
        salaries_city_level (dict[str: tuple[int, int]]): Уровень зарплат по городам
        vacancies_city_count (dict[str: int]): Количество вакансий по городам
    """

    def __init__(self, vacancy_info):
        """Инициализация объекта Report

        """
        self.salaries_year_level = vacancy_info[0]
        self.vacancies_year_count = vacancy_info[1]
        self.selected_salary_year_level = vacancy_info[2]
        self.selected_vacancy_year_count = vacancy_info[3]
        self.salaries_city_level = vacancy_info[4]
        self.vacancies_city_count = vacancy_info[5]

    def generate_excel(self, vacancy_name):
        """Создание excel-файла

        Args:
            vacancy_name (str): Название выбранной вакансии
        """
        workbook = Workbook()
        stats_by_year = workbook.worksheets[0]
        stats_by_year.title = "Cтатистика по годам"
        stats_by_city = workbook.create_sheet("Cтатистика по городам")

        stats_by_year.append(["Год", "Средняя зарплата", f"Средняя зарплата - {vacancy_name}",
                              "Количество вакансий", f"Количество вакансий - {vacancy_name}"])
        for i, year in enumerate(self.salaries_year_level.keys(), 2):
            stats_by_year.cell(row=i, column=1, value=year)
            for j, dictionary in enumerate((self.salaries_year_level, self.vacancies_year_count,
                                            self.selected_salary_year_level, self.selected_vacancy_year_count), 2):
                stats_by_year.cell(row=i, column=j, value=dictionary[year])

        stats_by_city.append(["Город", "Уровень зарплат", "", "Город", "Доля вакансий"])
        for i, city in enumerate(self.salaries_city_level.keys(), 2):
            stats_by_city.cell(row=i, column=1, value=city)
            stats_by_city.cell(row=i, column=2, value=self.salaries_city_level[city])
        for i, city in enumerate(self.vacancies_city_count.keys(), 2):
            stats_by_city.cell(row=i, column=4, value=city)
            stats_by_city.cell(row=i, column=5, value=self.vacancies_city_count[city])

        self.workbook(workbook)
        workbook.save('report.xlsx')

    @staticmethod
    def workbook(wb):
        """Стилизация excel-файла

        Args:
            wb (Workbook): Excel-лист
        """
        bold_font = Font(bold=True)
        thin = Side(border_style="thin", color="000000")
        outline = Border(top=thin, left=thin, right=thin, bottom=thin)
        for worksheet in wb.worksheets:
            for column_cells in worksheet.columns:
                length = max(len(str(cell.value) if cell.value is not None else "") for cell in column_cells)
                worksheet.column_dimensions[column_cells[0].column_letter].width = length + 3
            for cell in worksheet[1]:
                cell.font = bold_font
            for column in tuple(worksheet.columns):
                if column[1].value is None:
                    continue
                for cell in column:
                    cell.border = outline

    def generate_image(self, vacancy_name):
        """Создание графиков

        """
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(12, 7.5), layout='constrained')
        self.generate_salary_year_levels_graph(ax1, vacancy_name)
        self.generate_vacancy_year_count_graph(ax2, vacancy_name)
        self.generate_salary_city_levels_graph(ax3)
        self.generate_vacancy_city_count_graph(ax4)
        plt.savefig('graph.png')

    def generate_salary_year_levels_graph(self, ax, vacancy_name):
        """Создание графика уровня зарплат по годам

        Args:
            ax (Ax): Объект графика
            vacancy_name (str): Название выбранной вакансии
        """
        ax_labels = self.salaries_year_level.keys()
        x = np.arange(len(ax_labels))
        width = 0.35
        ax.bar(x - width / 2, self.salaries_year_level.values(), width, label='Средняя з/п')
        ax.bar(x + width / 2, self.selected_salary_year_level.values(), width, label=f'З/п {vacancy_name}')
        ax.set_xticks(x, ax_labels, fontsize=8, rotation=90, ha='right')
        ax.set_title("Уровень зарплат по годам")
        ax.yaxis.grid(True)
        ax.legend(fontsize=8, loc='upper left')

    def generate_vacancy_year_count_graph(self, ax, vacancy_name):
        """Создание графика количества вакансий по годам

        Args:
            ax (Ax): Объект графика
            vacancy_name (str): Название выбранной вакансии
        """
        ax_labels = self.vacancies_year_count.keys()
        x = np.arange(len(ax_labels))
        width = 0.35
        ax.bar(x - width / 2, self.vacancies_year_count.values(), width, label='Количество вакансий')
        ax.bar(x + width / 2, self.selected_vacancy_year_count.values(), label=f'Количество вакансий {vacancy_name}')
        ax.set_xticks(x, ax_labels, fontsize=8, rotation=90, ha='right')
        ax.set_title("Количество вакансий по годам")
        ax.yaxis.grid(True)
        ax.legend(fontsize=8, loc='upper left')

    def generate_salary_city_levels_graph(self, ax):
        """Создание графика уровня зарплат по городам

        Args:
            ax (Ax): Объект графика
        """
        ax_labels = self.salaries_city_level.keys()
        y_pos = np.arange(len(ax_labels))
        ax.barh(y_pos, self.salaries_city_level.values(), align='center')
        ax.set_yticks(y_pos, fontsize=8, labels=ax_labels)
        ax.invert_yaxis()
        ax.set_title("Уровень зарплат по городам")

    def generate_vacancy_city_count_graph(self, ax):
        """Создание графика количества вакансий по городам

        Args:
            ax (Ax): Объект графика
        """
        ax_labels, values = list(self.vacancies_city_count.keys()), self.vacancies_city_count.values()
        ax_labels.append('Другие')
        values = list(map(lambda value: float(value[:-1]), values))
        values.append(100 - sum(values))
        ax.pie(values, labels=ax_labels)
        ax.set_title("Доля вакансий по городам")

    def generate_pdf(self, vacancy_name):
        """Создание pdf-файла

        """
        headers1, headers2, headers3 = (["Год", "Средняя зарплата", f"Средняя зарплата - {vacancy_name}",
                                         "Количество вакансий", f"Количество вакансий - {vacancy_name}"],
                                        ["Город", "Уровень зарплат"], ["Город", "Доля вакансий"])
        rows1 = list(map(lambda year: [year] + [dictionary[year] for dictionary in
                                                (self.salaries_year_level, self.vacancies_year_count,
                                                 self.selected_salary_year_level, self.selected_vacancy_year_count)]
                         , self.salaries_year_level.keys()))
        rows2 = list(map(lambda city: [city, self.salaries_city_level[city]], self.salaries_city_level.keys()))
        rows3 = list(map(lambda city: [city, self.vacancies_city_count[city]], self.vacancies_city_count.keys()))

        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template.html")
        pdf_template = template.render(graph_name='graph.png',
                                       vacancy_name=vacancy_name, headers1=headers1, headers2=headers2,
                                       headers3=headers3, rows1=rows1, rows2=rows2, rows3=rows3)
        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        options = {'enable-local-file-access': None}
        pdfkit.from_string(pdf_template, 'report.pdf', options=options, configuration=config)


######################################################################################################################


def get_statistics():
    """Получение информации с csv файла и создание графиков, таблиц и общего pdf-файл со статистикой
        на основе вводимых пользователем данных

    """
    input_requests = ["Введите название файла: ", "Введите название профессии: "]
    input_info = [input(input_request) for input_request in input_requests]
    # input_info = ["vacancies_by_year.csv", "Javascript"]
    if os.stat(input_info[0]).st_size == 0:
        print("Пустой файл")
        return
    data_set = DataSet(input_info[0])
    if len(data_set.vacancies_objects) == 0:
        print("Нет данных")
        return
    input_connect = InputConnect()
    formatted_info = input_connect.info_formatter(data_set.vacancies_objects)
    info = input_connect.info_finder(formatted_info, input_info[1])
    report = Report(info)
    report.generate_excel(input_info[1])
    report.generate_image(input_info[1])
    report.generate_pdf(input_info[1])
